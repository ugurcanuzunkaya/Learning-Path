from itertools import product
from math import sqrt
import numpy as np
import gurobipy as gp
from gurobipy import GRB
import time
import random
from openpyxl import load_workbook
import openpyxl
import pandas as pd

start = time.time()

wb = openpyxl.Workbook()
Sheet_name = wb.sheetnames
distance = pd.read_csv("distances.csv", header=None, sep=';')
distance = distance.iloc[0:, 0:].to_numpy()
for i in range(len(distance)):
    for j in range(len(distance)):
        if i==j:
            distance[i,j] = 99999999999

ii = len(distance)
demand = [0, 6, 10, 7, 7, 7, 7, 1, 9, 2, 8,	7, 9, 9, 9]


k = 6
cap = 25
obj = np.zeros([10,4])
w1 = 1

ideal1 = 185
ideal2 = 36
while True:
    w2 = 11 - w1
    set_I = range(1, ii+1)
    set_J = range(1, ii+1)
    set_K = range(1, k+1)

    m1 = gp.Model('OVRP')
    m1.setParam(GRB.Param.IntegralityFocus, 1)
    x_ijk = m1.addVars(set_I, set_J, set_K,  vtype=GRB.BINARY, name='x_ijk')
    y_ik = m1.addVars(set_I, set_K, vtype=GRB.BINARY, name='y_ik')
    y_jk = m1.addVars(set_J, set_K, vtype=GRB.BINARY, name='y_jk')
    a_i = m1.addVars(set_I, vtype=GRB.CONTINUOUS, lb=0.0, name='a_i')
    a_j = m1.addVars(set_J, vtype=GRB.CONTINUOUS, lb=0.0, name='a_j')
    beta = m1.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='Beta')
    t = m1.addVar(vtype=GRB.CONTINUOUS, lb=0.0, name='t')

    m1.update()

    m1.addConstrs(y_jk[j,k] - y_ik[i,k] == 0 for i in set_I for j in set_J if i == j for k in set_K)
    m1.addConstrs(a_j[j] - a_i[i] == 0 for i in set_I for j in set_J if i == j)

    m1.addConstrs((gp.quicksum(distance[i-1, j-1]*x_ijk[i,j,k] for i in set_I for j in set_J) <= beta for k in set_K), name='secondOBJ')
    m1.addConstrs((gp.quicksum(demand[i-1]*y_ik[i, k] for i in set_I) <= cap for k in set_K), name='capacity')

    m1.addConstrs((gp.quicksum(y_ik[i, k] for k in set_K) == 1 for i in set_I if i != 1), name='assignment')

    m1.addConstrs((gp.quicksum(x_ijk[i, j, k] for i in set_I) == y_jk[j,k] for j in set_J if j != 1 for k in set_K ), name='relation')

    m1.addConstrs((gp.quicksum(x_ijk[i, j, k] for j in set_J if j != 1) <= y_ik[i, k] for i in set_I for k in set_K), name='r2')
    m1.addConstrs((y_ik[i,k] == 1 for i in set_I for k in set_K if i == 0), name='startpoint')
    m1.addConstrs(((gp.quicksum(x_ijk[i, j, k] for k in set_K))*ii + a_i[i] - a_j[j] <= ii -1 for i in set_I for j in set_J if i != j if i != 1 if j != 1), name='subtour')
    m1.addConstr((w1*(gp.quicksum(distance[i-1, j-1]*x_ijk[i,j,k] for i in set_I for j in set_J for k in set_K) - ideal1) <= t), name='chebcons1')
    m1.addConstr((w2 * (beta - ideal2) <= t), name='chebcons2')
    m1.setObjective(t, GRB.MINIMIZE)
    m1.write('model_obj5.lp')
    m1.optimize()

    total_distance = 0
    if m1.status not in [GRB.INFEASIBLE, GRB.UNBOUNDED, GRB.INF_OR_UNBD]:
        for i in set_I:
            for j in set_J:
                for k in set_K:
                    if x_ijk[i, j, k].x != 0:
                        total_distance = total_distance + distance[i-1][j-1]

        obj[w1 - 1][0] = total_distance
        obj[w1 - 1][1] = beta.x
        obj[w1 - 1][2] = w1
        obj[w1 - 1][3] = w2
        print('******************************')
        print(m1.objVal)
    else:
        obj[w1 - 1][0] = 0
        obj[w1 - 1][1] = 0
        obj[w1 - 1][2] = w1
        obj[w1 - 1][3] = w2
    w1 = w1+1
    if w1 == 11:
        break
print(obj)








